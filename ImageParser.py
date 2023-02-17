import requests
import openpyxl
from tqdm import tqdm
import concurrent.futures
from PIL import Image
from io import BytesIO
import config


url = 'https://www.sima-land.ru/'
api_token = config.api_token
headers = {'x-api-key': api_token,
           'Accept': 'application/json'}
session = requests.Session()
session.headers = headers
TIMEOUT = 10


def get_article_list(filename=config.articles_file):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    articles = []
    for cell in sheet['A']:
        articles.append(cell.value)
    return articles


def get_product_info(article):
    for i in range(len(str(article))):
        if article[i].isdigit():
            product_article = article[i:]
            break
    url = f'https://www.sima-land.ru/api/v3/item/?sid={product_article}'
    resp = session.get(url, headers=headers, timeout=TIMEOUT)
    data = resp.json()
    product_name = data['items'][0]['name']
    photo_url = data['items'][0]['photoUrl']
    product = {'article': product_article,
               'name': product_name,
               'image': photo_url}
    return product


def get_img(url, size=(100, 100)):
    r = session.get(url, stream=True)
    if not r.ok:
        print('Image', url, ' ---ERROR')
    r.raw.decode_content = True
    img = Image.open(r.raw)
    if size:
        img = img.resize(size)
    temp = BytesIO()
    img.save(temp, format="png")
    temp.seek(0)
    return Image.open(temp)


def create_table(products):
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Sima Land', index=0)
    ws = wb['Sima Land']
    ws[f'A1'] = 'Фото'
    ws[f'B1'] = 'Артикул'
    ws[f'C1'] = 'Название товара'
    i = 2
    for product in products:
        photo = openpyxl.drawing.image.Image(get_img(product['image'], size=(150, 150)))
        ws.row_dimensions[i].height = int(150 * .8)
        ws.column_dimensions["A"].width = int(150 * .2)
        ws.column_dimensions["C"].width = 70
        ws[f'B{i}'] = product['article']
        ws[f'C{i}'] = product['name']
        ws.add_image(photo, f'A{i}')
        i += 1
    wb.save(f'ImagesResults.xlsx')


def get_all_products():
    articles = get_article_list()
    all_products = []
    CONNECTIONS = 8
    with concurrent.futures.ThreadPoolExecutor(max_workers=CONNECTIONS) as executor:
        future_to_article = (executor.submit(get_product_info, article) for article in articles)
        for future in tqdm(concurrent.futures.as_completed(future_to_article), total=len(articles)):
            try:
                product = future.result()
            except Exception as exc:
                data = str(type(exc))
                print(data)
            finally:
                all_products.append(product)
    return all_products


def parse():
    all_products = get_all_products()
    create_table(all_products)


if __name__ == '__main__':
    parse()
