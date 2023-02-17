import requests
import openpyxl
from tqdm import tqdm
import concurrent.futures
import config


url = 'https://www.sima-land.ru/'

api_token = config.api_token
headers = {'x-api-key': api_token,
           'Accept': 'application/json'}
session = requests.Session()
session.headers = headers
TIMEOUT = 10


def get_article_list(filename=config.articles_file):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    articles = []
    for cell in sheet['A']:
        articles.append(cell.value)
    return articles


def get_product_info(article):
    for i in range(len(str(article))):
        if article[i].isdigit():
            product_article = article[i:]
            break
    url = f'https://www.sima-land.ru/api/v3/item/?sid={product_article}'
    resp = session.get(url, headers=headers, timeout=TIMEOUT)
    data = resp.json()
    try:
        product_balance = data['items'][0]['balance']
        product_partner = 'Товар партнёра' if data['items'][0]['is_remote_store'] else ''
        product = {'article': article,
                   'balance': product_balance,
                   'partner': product_partner}
    except Exception as ex:
        print(ex)
        product = {'article': article,
                   'balance': 0,
                   'partner': 'Товара нет на сайте'}
    return product


def create_table(products):
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Sima Land', index=0)
    ws = wb['Sima Land']
    ws[f'A1'] = 'Артикул'
    ws[f'B1'] = 'Остаток'
    ws[f'C1'] = 'Товар партнера'
    i = 2
    for product in products:
        ws[f'A{i}'] = product['article']
        ws[f'B{i}'] = product['balance']
        ws[f'C{i}'] = product['partner']
        i += 1
    wb.save(f'Results.xlsx')


def get_all_products():
    articles = get_article_list()
    all_products = []
    CONNECTIONS = 8

    with concurrent.futures.ThreadPoolExecutor(max_workers=CONNECTIONS) as executor:
        future_to_article = (executor.submit(get_product_info, article) for article in articles)
        for future in tqdm(concurrent.futures.as_completed(future_to_article), total=len(articles)):
            try:
                product = future.result()
            except Exception as exc:
                data = str(type(exc))
                print(data)
            finally:
                all_products.append(product)
    return all_products


def filter_products(all_products):
    products = []
    for product in all_products:
        if product['balance'] < 5 or product['partner'] == 'Товар партнёра':
            products.append(product)
    return products


def parse():
    all_products = get_all_products()
    products = filter_products(all_products)
    create_table(products)


if __name__ == '__main__':
    parse()
